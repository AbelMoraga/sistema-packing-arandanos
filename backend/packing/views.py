from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.db.models import Q, Sum, Count, DecimalField, Value, F
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models.functions import Coalesce,TruncDate
from django.utils import timezone
from .forms import PalletForm
from datetime import timedelta
from django.template.loader import render_to_string
from openpyxl.utils import get_column_letter
import csv
from django.contrib.auth.models import User, Group
from openpyxl import Workbook




from .models import (
    
    Pallet,
    GrupoProceso,
    Distribuidor,
    TipoPocillo,
    Variedad,
    Linea,
    PalletTerminado,
    IQFDescarte,
    Productor,
    TipoEnvase,
    RegistroActividad,
)




def es_admin_local(user):
    return user.groups.filter(name="admin_local").exists() or user.is_superuser

@user_passes_test(es_admin_local)
def panel_admin(request):
    return render(request, "packing/admin/panel_admin.html")



def es_frio_o_control(user):
    return (
        user.groups.filter(name="Frio").exists()
        or user.groups.filter(name="Control").exists()
        or user.is_superuser
    )

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        print(f"Intentando login con: {username}")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            registrar_actividad(
                request,
                accion="Inicio de sesión",
                modulo="Autenticación",
                detalle=f"Usuario {user.username} inició sesión"
            )


            print(f"Autenticado como: {user.username}")
            print(f"Grupos: {user.groups.all()}")

            if user.groups.filter(name="Frio").exists():
                messages.success(request, "Bienvenido al área de Frío ❄️")
                return redirect("menu_frio")
            elif user.groups.filter(name="Procesos").exists():
                messages.success(request, "Bienvenido al área de Procesos ⚙️")
                return redirect("menu_procesos")
            elif user.groups.filter(name="Control").exists():
                messages.success(request, "Bienvenido al área de Control 📊")
                return redirect("menu_control")
            elif user.groups.filter(name="Admin_local").exists():
                return redirect("panel_admin")  
            else:
                messages.error(request, "No tienes un rol asignado.")
                return redirect("login")
        else:
            messages.error(request, "Usuario o contraseña incorrectos ❌")
            return redirect("login")

    return render(request, "packing/base.html")


# Rp
def registrar_pallets_entrada(request):
    success = False  # bandera

    if request.method == 'POST':
        form = PalletForm(request.POST)
        if form.is_valid():
            form.save()

            registrar_actividad(
                request,
                accion="Crear pallet",
                modulo="Recepción",
                detalle=f"Pallet ID {pallet.id} creado en recepción"
            )
            
            success = True  # activamos notificación
            form = PalletForm()  # limpiar el formulario después de guardar
    else:
        form = PalletForm()

    return render(request, 'packing/registrar_pallets_entrada.html', {'form': form, 'success': success})

def success(request):
    return render(request, 'packing/success.html')

def lista_pallets(request):
    pallets = Pallet.objects.all().order_by('bloqueado_recepcion', '-id')
    return render(request, 'packing/lista_pallets.html', {'pallets': pallets})

#ACTUALIZAR ESTADOS
def actualizar_estado(request, id):
    pallet = get_object_or_404(Pallet, id=id)
    accion = request.GET.get('accion')
    origen = request.GET.get('origen', 'pallets')


    if origen == 'pallets' and not pallet.bloqueado_recepcion:
        if accion == 'enfriado':
            pallet.enfriado = not pallet.enfriado
        elif accion == 'fumigado' and not pallet.organico:
            pallet.fumigado = not pallet.fumigado
        elif accion == 'preprocesos':
            pallet.pre_procesos = not pallet.pre_procesos
        elif accion == 'organico':
            pallet.organico = True
            pallet.fumigado = False
        elif accion == 'guardar':
            pallet.bloqueado_recepcion = True  
        pallet.save()

        registrar_actividad(
            request,
            accion="Actualizar estado pallet",
            modulo="Recepción",
            detalle=f"Pallet ID {pallet.id} | Acción: {accion}"
        )
        
        return redirect('lista_pallets')

    #PROCESOS
    elif origen == 'procesos':
        if not pallet.bloqueado_proceso:
            if accion == 'procesado':
                pallet.procesado = not pallet.procesado
                pallet.save()
                return redirect(request.META.get('HTTP_REFERER', 'lista_procesos'))
            elif accion == 'guardar_proceso':
                if pallet.procesado:
                    pallet.bloqueado_proceso = True  
                    pallet.save()
                    return redirect('lista_procesos')

    return redirect('lista_procesos' if origen == 'procesos' else 'lista_pallets')

#procesos
def lista_procesos(request):
    pallets = Pallet.objects.filter(pre_procesos=True).order_by('bloqueado_proceso', 'procesado', '-id')
    buscar_id = request.GET.get("buscar_id")
    if buscar_id:
        pallets = pallets.filter(id=buscar_id)

    return render(request, 'packing/lista_procesos.html', {'pallets': pallets})

def crear_grupo_proceso(request):
    if request.method == 'POST':
        seleccionados = request.POST.getlist('pallets_seleccionados')

        if not seleccionados:
            messages.error(request, "Debes seleccionar al menos un pallet.")
            return redirect('lista_procesos')

        pallets = Pallet.objects.filter(id__in=seleccionados)
        if any(p.grupo for p in pallets):
            messages.error(request, "Uno o más pallets ya pertenecen a un grupo existente.")
            return redirect('lista_procesos')

        productores = set(p.productor for p in pallets)
        variedades = set(p.variedad for p in pallets)

        if len(productores) > 1 or len(variedades) > 1:
            messages.error(request, "Solo puedes agrupar pallets con el mismo productor y variedad.")
            return redirect('lista_procesos')

        grupo = GrupoProceso.objects.create(
            productor=pallets[0].productor,
            variedad=pallets[0].variedad
        )

        for p in pallets:
            p.grupo = grupo
            p.save()

            registrar_actividad(
            request,
            accion="Crear grupo de proceso",
            modulo="Procesos",
            detalle=f"Grupo {grupo.id_grupo} con {len(pallets)} pallets"
        )


        messages.success(request, f"✅ Grupo {grupo.id_grupo} creado correctamente con {len(pallets)} pallets.")
        return redirect('lista_procesos')

    return redirect('lista_procesos')


def menu_frio(request):
    return render(request, 'packing/menu_frio.html')

def menu_procesos(request):
    return render(request, 'packing/menu_procesos.html')

def logout_view(request):

   
    registrar_actividad(
        request,
        accion="Cierre de sesión",
        modulo="Autenticación",
        detalle=f"Usuario {request.user.username} cerró sesión"
    )

    logout(request)
    messages.success(request, "Sesión cerrada correctamente 👋")
    return redirect("login")


def pallet_terminado(request):
    distribuidores = Distribuidor.objects.all()
    tipos_pocillo = TipoPocillo.objects.all()
    lineas = Linea.objects.all()
    variedades = Variedad.objects.all()  
    grupos = GrupoProceso.objects.all()

    if request.method == 'POST':
        distribuidor_id = request.POST.get('distribuidor')
        tipo_pocillo_id = request.POST.get('tipo_pocillo')
        linea_id = request.POST.get('linea')
        variedad_id = request.POST.get('variedad')  
        cantidad = request.POST.get('cantidad')
        calidad_color = request.POST.get('calidad')

    
        if not (distribuidor_id and tipo_pocillo_id and linea_id and variedad_id and cantidad and calidad_color):
            messages.error(request, "⚠️ Todos los campos son obligatorios.")
            return render(request, 'packing/pallet_terminado.html', {
                'distribuidores': distribuidores,
                'tipos_pocillo': tipos_pocillo,
                'lineas': lineas,
                'variedades': variedades,
                'grupos': grupos,
            })

        
        calidad_dict = {
            'rojo': 'Rechazado',
            'amarillo': 'Baja',
            'verde': 'Buena',
            'azul': 'Excelente'
        }
        calidad = calidad_dict.get(calidad_color, 'Sin definir')

        
        PalletTerminado.objects.create(
            distribuidor_id=distribuidor_id,
            tipo_pocillo_id=tipo_pocillo_id,
            linea_id=linea_id,
            variedad_id=variedad_id,  
            cantidad_cajas=cantidad,
            calidad=calidad
        )

        messages.success(request, f"✅ Pallet registrado con calidad: {calidad}")
        return redirect('pallet_terminado')

    return render(request, 'packing/pallet_terminado.html', {
        'distribuidores': distribuidores,
        'tipos_pocillo': tipos_pocillo,
        'lineas': lineas,
        'variedades': variedades,  
        'grupos': grupos,
    })



def registrar_iqf_descarte(request):
 
    grupos_disponibles = GrupoProceso.objects.filter(iqfdescarte__isnull=True)

    if request.method == 'POST':
        print("📦 Datos POST recibidos:", request.POST)

        grupo_id = request.POST.get('grupo_proceso', '').strip()
        peso_iqf = request.POST.get('peso_iqf', '').strip()
        peso_descarte = request.POST.get('peso_descarte', '').strip()

  
        if not grupo_id or not peso_iqf or not peso_descarte:
            messages.error(request, "⚠️ Completa todos los campos antes de guardar.")
            return render(request, 'packing/iqf.html', {'grupos': grupos_disponibles})

        try:
            grupo = GrupoProceso.objects.get(id_grupo=grupo_id)

          
            peso_iqf = float(peso_iqf.replace(',', '.'))
            peso_descarte = float(peso_descarte.replace(',', '.'))

           
            IQFDescarte.objects.create(
                
                grupo_proceso=grupo,
                peso_iqf=peso_iqf,
                peso_descarte=peso_descarte,
                productor=str(grupo.productor),    
                variedad=str(grupo.variedad),      
                estado="EN_IQF"                 

            )
            
            
            registrar_actividad(
                request,
                accion="Registrar IQF/Descarte",
                modulo="Procesos",
                detalle=f"Grupo {grupo.id_grupo} | IQF: {peso_iqf} kg | Descarte: {peso_descarte} kg"
            )

            messages.success(
                request,
                f"✅ IQF y Descarte registrados correctamente para el Grupo {grupo.id_grupo}"
            )

      
            grupos_disponibles = GrupoProceso.objects.filter(iqfdescarte__isnull=True)

            return render(
                request,
                'packing/iqf.html',
                {
                    'grupos': grupos_disponibles,
                    'peso_iqf': '',
                    'peso_descarte': ''
                }
            )

        except ValueError:
            messages.error(request, "⚠️ Los pesos deben ser números válidos (usa punto o coma).")

        except GrupoProceso.DoesNotExist:
            messages.error(request, "❌ El grupo seleccionado no existe.")

    return render(request, 'packing/iqf.html', {'grupos': grupos_disponibles})



def lista_pallets_terminados(request):
    pallets = PalletTerminado.objects.all().order_by('fecha')

    estado = request.GET.get('estado')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    buscar = request.GET.get('buscar', '').strip()

    if estado and estado != 'Todos':
        pallets = pallets.filter(estado__iexact=estado)

    if fecha_inicio:
        pallets = pallets.filter(fecha__date__gte=parse_date(fecha_inicio))

    if fecha_fin:
        pallets = pallets.filter(fecha__date__lte=parse_date(fecha_fin))

    if buscar:
        pallets = pallets.filter(
            Q(id__icontains=buscar) |
            Q(distribuidor__nombre__icontains=buscar)
        )

    cambiar_id = request.GET.get('cambiar_id')
    if cambiar_id:
        p = PalletTerminado.objects.get(id=cambiar_id)
        if p.estado == 'Por enviar':
            p.estado = 'Enviado'
            p.save()

            registrar_actividad(
                request,
                accion="Enviar pallet terminado",
                modulo="Procesos",
                detalle=f"PalletTerminado ID {p.id} enviado a despacho"
            )
            
        return redirect('lista_pallets_terminados')

    distribuidores = Distribuidor.objects.all()

    return render(request, 'packing/lista_pallets_terminados.html', {
        'pallets': pallets,
        'distribuidores': distribuidores,
    })

def es_control(user):
    """Verifica si el usuario pertenece al grupo 'Control'."""
    return user.groups.filter(name="Control").exists()

@login_required
@user_passes_test(es_control)
def menu_control(request):
    """Muestra el menú principal del área de Control."""
    return render(request, 'packing/menu_control.html')


#  GERENCIA 


@login_required
@user_passes_test(es_control)
def control_iqf(request):
    return render(request, 'packing/control/iqf.html')

@login_required
@user_passes_test(es_control)
def control_pallets(request):
    return render(request, 'packing/control/pallets.html')



#reportes
@login_required
@user_passes_test(es_control)
def control_reportes(request):

    total_pallets = Pallet.objects.count()
    total_procesados = Pallet.objects.filter(procesado=True).count()
    total_terminados = PalletTerminado.objects.count()

    total_iqf = IQFDescarte.objects.aggregate(
        total=Sum("peso_iqf")
    )["total"] or 0

    total_descarte = IQFDescarte.objects.aggregate(
        total=Sum("peso_descarte")
    )["total"] or 0

    calidad_terminados = (
        PalletTerminado.objects
        .values("calidad")
        .annotate(total=Count("id"))
    )

    pallets_por_productor = (
        Pallet.objects
        .values("productor__nombre")
        .annotate(total=Count("id"))
        .order_by("productor__nombre")
    )

    return render(request, "packing/control/reportes.html", {
        "total_pallets": total_pallets,
        "total_procesados": total_procesados,
        "total_terminados": total_terminados,
        "total_iqf": total_iqf,
        "total_descarte": total_descarte,
        "calidad_terminados": calidad_terminados,
        "pallets_por_productor": pallets_por_productor,
    })


#parte compleja---------------------------------------------------------------------------------------------------
"""
@login_required
@user_passes_test(es_control)
def lista_iqf(request):

    iqf = IQFDescarte.objects.all().order_by('-id')


    grupo = request.GET.get("grupo")
    if grupo:
        iqf = iqf.filter(grupo_proceso__id_grupo__icontains=grupo)

    estado = request.GET.get("estado")
    if estado == "en_iqf":
        iqf = iqf.filter(estado="EN_IQF")
    elif estado == "retirado":
        iqf = iqf.filter(estado="RETIRADO")

    variedad = request.GET.get("variedad")
    if variedad and variedad != "todas":
        iqf = iqf.filter(variedad__iexact=variedad)

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    if fecha_inicio and fecha_fin:
        iqf = iqf.filter(fecha_ingreso__date__range=[fecha_inicio, fecha_fin])


    riesgo_param = request.GET.get("riesgo")
    hace_10_dias = timezone.now() - timedelta(days=10)

    if riesgo_param == "true":
        iqf = iqf.filter(estado="EN_IQF", fecha_ingreso__lte=hace_10_dias)


    pallets_iqf = IQFDescarte.objects.filter(estado="EN_IQF").count()
    retirados = IQFDescarte.objects.filter(estado="RETIRADO").count()
    riesgo = IQFDescarte.objects.filter(estado="EN_IQF", fecha_ingreso__lte=hace_10_dias).count()


    variedades = IQFDescarte.objects.values_list("variedad", flat=True).distinct()

    return render(request, 'packing/control/iqf_partial.html', {
        'iqf': iqf,
        'pallets_iqf': pallets_iqf,
        'retirados': retirados,
        'riesgo': riesgo,
        'variedades': variedades,
    })

"""

# DETALLES

@login_required
@user_passes_test(es_control)
def detalles_iqf(request, id):

    registro = get_object_or_404(IQFDescarte, id=id)
    context = {"i": registro}
    template = "packing/control/detalles_iqf_partial.html"

  
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return render(request, template, context)

    contenido = render_to_string(template, context, request=request)

    return render(request, "packing/menu_control.html", {
        "contenido_inicial": contenido
    })

## terminados 


@login_required
@user_passes_test(es_control)
def control_pallets_terminados(request):

    pallets = PalletTerminado.objects.all().order_by("-id")

    buscar = request.GET.get("buscar")
    if buscar:
        pallets = pallets.filter(
            Q(id__icontains=buscar) |
            Q(distribuidor__nombre__icontains=buscar)
        )

    estado = request.GET.get("estado")
    if estado:
        pallets = pallets.filter(estado=estado)

    variedad = request.GET.get("variedad")
    if variedad:
        pallets = pallets.filter(variedad_id=variedad)

    context = {
        "pallets": pallets,
        "total": pallets.count(),
        "enviados": pallets.filter(estado="Enviado").count(),
        "por_enviar": pallets.filter(estado="Por enviar").count(),
        "variedades": Variedad.objects.all(),
    }

    return render(request, "packing/control/pallets_terminados.html", context)



@login_required
@user_passes_test(es_control)
def control_exportar(request):
    return render(request, "packing/control/exportar.html")
#exportar


@login_required
@user_passes_test(es_control)
def exportar_pallets_terminados(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pallets Terminados"


    headers = ["ID", "Distribuidor", "Variedad", "Cantidad", "Calidad", "Fecha", "Estado"]
    ws.append(headers)


    for p in PalletTerminado.objects.all():
        ws.append([
            p.id,
            p.distribuidor.nombre if p.distribuidor else "—",
            p.variedad.nombre if p.variedad else "—",
            p.cantidad_cajas,
            p.calidad,
            p.fecha.strftime("%d-%m-%Y") if p.fecha else "—",
            p.estado,
        ])


    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="pallets_terminados.xlsx"'

    wb.save(response)
    return response


# EXPORTAR DATOS 
import openpyxl
from openpyxl.utils import get_column_letter
import csv


@login_required
@user_passes_test(es_control)
def exportar_recepcion_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recepción de Pallets"

    ws.append([
        "ID",
        "Código",
        "Productor",
        "Variedad",
        "Tipo Envase",
        "Cantidad Cajas",
        "Peso Neto",
        "Fecha Recepción",
    ])

    pallets = Pallet.objects.all().order_by("-fecha_creacion")

    for p in pallets:
        ws.append([
            p.id,
            p.codigo,
            p.productor.nombre,
            p.variedad.nombre,
            p.tipo_envase.nombre,
            p.cantidad_cajas,
            float(p.peso_neto),
            p.fecha_creacion.strftime("%d-%m-%Y %H:%M"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="recepcion_pallets.xlsx"'
    wb.save(response)

    return response


@login_required
@user_passes_test(es_control)
def exportar_procesos_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="procesos.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "ID Grupo",
        "Productor",
        "Variedad",
        "Fecha creación",
        "Cantidad pallets",
        "Pallets procesados",
    ])

    grupos = GrupoProceso.objects.all().order_by("-fecha_creacion")

    for g in grupos:
        pallets = Pallet.objects.filter(grupo=g)
        procesados = pallets.filter(procesado=True).count()

        writer.writerow([
            g.id_grupo,
            g.productor.nombre,
            g.variedad.nombre,
            g.fecha_creacion.strftime("%d-%m-%Y %H:%M"),
            pallets.count(),
            procesados,
        ])

    return response


@login_required
@user_passes_test(es_control)
def exportar_pallets_terminados_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pallets Terminados"

    ws.append(["ID", "Distribuidor", "Variedad", "Cajas", "Calidad", "Estado"])

    for p in PalletTerminado.objects.all():
        ws.append([
            p.id,
            str(p.distribuidor),
            str(p.variedad),
            p.cantidad_cajas,
            p.calidad,
            p.estado,
        ])

    # Auditoría
    RegistroActividad.objects.create(
        usuario=request.user,
        accion="Exportar pallets terminados",
        modulo="Control",
        detalle="Exportación Excel de pallets terminados"
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pallets_terminados.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(es_control)
def exportar_iqf_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "IQF y Descarte"

    ws.append([
        "ID", "Grupo", "Productor", "Variedad",
        "Peso IQF", "Peso Descarte", "Estado"
    ])

    for i in IQFDescarte.objects.all():
        ws.append([
            i.id,
            i.grupo_proceso.id_grupo,
            str(i.productor),
            i.variedad,
            i.peso_iqf,
            i.peso_descarte,
            i.estado,
        ])

    RegistroActividad.objects.create(
        usuario=request.user,
        accion="Exportar IQF",
        modulo="Control",
        detalle="Exportación Excel de IQF y descarte"
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="iqf_descarte.xlsx"'
    wb.save(response)
    return response
    # Crear archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=\"iqf_descarte.xlsx\"'

    wb.save(response)
    return response


#volvamos el problema del meni

def control_iqf_lista(request):

    iqf = IQFDescarte.objects.all().order_by("-id")

    buscar = request.GET.get("buscar", "").strip()
    if buscar:
        iqf = iqf.filter(
            Q(id__icontains=buscar) |
            Q(grupo_proceso__id_grupo__icontains=buscar)
        )

    estado = request.GET.get("estado", "")
    if estado:
        iqf = iqf.filter(estado=estado)

    variedad = request.GET.get("variedad", "")
    if variedad:
        iqf = iqf.filter(variedad__iexact=variedad)

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    if fecha_inicio and fecha_fin:
        iqf = iqf.filter(fecha_ingreso__date__range=[fecha_inicio, fecha_fin])

    variedades = IQFDescarte.objects.values_list("variedad", flat=True).distinct()

    return render(request, "packing/control/iqf_lista.html", {
        "iqf": iqf,
        "variedades": variedades,
    })

def lista_iqf_full(request):

    iqf = IQFDescarte.objects.all().order_by("-id")

    buscar = request.GET.get("buscar", "").strip()
    if buscar:
        iqf = iqf.filter(
            Q(id__icontains=buscar) |
            Q(grupo_proceso__id_grupo__icontains=buscar)
        )

    estado = request.GET.get("estado")
    if estado:
        iqf = iqf.filter(estado=estado)

    variedad = request.GET.get("variedad")
    if variedad:
        iqf = iqf.filter(variedad__iexact=variedad)

    variedades = IQFDescarte.objects.values_list("variedad", flat=True).distinct()

    return render(request, "packing/control/lista_iqf_full.html", {
        "iqf": iqf,
        "variedades": variedades,
        "buscar": buscar,
        "estado_filtro": estado,
        "variedad_filtro": variedad,
    })

#adm


@user_passes_test(es_admin_local)
def panel_admin_usuarios(request):
    from django.contrib.auth.models import User, Group

    usuarios = User.objects.all()
    grupos = Group.objects.all()

    return render(request, "packing/admin/panel_admin_usuarios.html", {
        "usuarios": usuarios,
        "grupos": grupos
    })




@login_required
@user_passes_test(es_admin_local)
def panel_admin_editar(request, modelo, id):

    # MODELOS DISPONIBLES
    modelos = {
        "pallet": Pallet,
        "grupoproceso": GrupoProceso,
        "iqf": IQFDescarte,
        "terminados": PalletTerminado,
    }

    ModelClass = modelos.get(modelo)
    if not ModelClass:
        return HttpResponse("Modelo no válido")

    # ===================================================
    #                 LISTADO + BUSCADOR
    # ===================================================
    if id == 0:
        objetos = ModelClass.objects.all().order_by("-id")

        # ---- FILTRO BUSCAR (MISMO NOMBRE PARA TODOS) ----
        buscar = request.GET.get("buscar", "").strip()
        if buscar:
            if modelo == "pallet":
                # Busca por ID o código
                objetos = objetos.filter(
                    Q(id__icontains=buscar) |
                    Q(codigo__icontains=buscar)
                )
            else:
                # Para terminados / iqf / otros → solo por ID
                objetos = objetos.filter(id__icontains=buscar)

        # Selección de template LISTA según modelo
        if modelo == "terminados":
            template = "packing/admin/panel_admin_lista_terminados.html"
        elif modelo == "pallet":
            template = "packing/admin/panel_admin_lista_pallet.html"
        elif modelo == "iqf":
            template = "packing/admin/panel_admin_lista_iqf.html"
        else:
            template = "packing/admin/panel_admin_lista.html"

        return render(request, template, {
            "modelo": modelo,
            "objetos": objetos,
            "buscar": buscar,
        })

    # ===================================================
    #                  DETALLE / EDITAR
    # ===================================================

    obj = get_object_or_404(ModelClass, id=id)

    if request.method == "POST":

        # ---------- EDITAR PALLET NORMAL ----------
        if modelo == "pallet":
            obj.productor_id = request.POST.get("productor")
            obj.variedad_id = request.POST.get("variedad")
            obj.tipo_envase_id = request.POST.get("tipo_envase")
            obj.cantidad_cajas = request.POST.get("cantidad_cajas")
            obj.peso_neto = request.POST.get("peso_neto")
            obj.save()
            registrar_actividad(
                request,
                accion="Editar pallet",
                modulo="Administración",
                detalle=f"Pallet ID {obj.id} editado por admin"
            )

            messages.success(request, "Cambios guardados correctamente")
            return redirect(f"/panel-admin/editar/{modelo}/{obj.id}/")

        # ---------- EDITAR PALLET TERMINADO ----------
        if modelo == "terminados":
            obj.distribuidor_id = request.POST.get("distribuidor")
            obj.variedad_id = request.POST.get("variedad")
            obj.cantidad_cajas = request.POST.get("cantidad_cajas")
            obj.calidad = request.POST.get("calidad")
            obj.estado = request.POST.get("estado")
            obj.save()

            registrar_actividad(
                request,
                accion="Editar pallet terminado",
                modulo="Administración",
                detalle=f"Pallet terminado ID {obj.id} | Estado: {obj.estado}"
            )

            messages.success(request, "Cambios guardados correctamente")
            return redirect(f"/panel-admin/editar/{modelo}/{obj.id}/")

        # ---------- EDITAR IQF / DESCARTE ----------
        if modelo == "iqf":
            # Pesos (aceptamos números, si algo viene vacío se ignora)
            peso_iqf = request.POST.get("peso_iqf", "").replace(",", ".")
            peso_descarte = request.POST.get("peso_descarte", "").replace(",", ".")
            peso_final = request.POST.get("peso_final", "").replace(",", ".")

            try:
                if peso_iqf != "":
                    obj.peso_iqf = float(peso_iqf)
            except ValueError:
                pass

            try:
                if peso_descarte != "":
                    obj.peso_descarte = float(peso_descarte)
            except ValueError:
                pass

            try:
                if peso_final != "":
                    obj.peso_final = float(peso_final)
            except ValueError:
                pass

            estado = request.POST.get("estado")
            if estado in ["EN_IQF", "RETIRADO"]:
                obj.estado = estado

            obj.save()
            registrar_actividad(
                request,
                accion="Editar IQF / Descarte",
                modulo="Administración",
                detalle=f"IQF ID {obj.id} | Estado: {obj.estado}"
            )

            messages.success(request, "Cambios guardados correctamente")
            return redirect(f"/panel-admin/editar/{modelo}/{obj.id}/")

        # ---------- OTROS MODELOS (grupo, etc.) ----------
        for campo, valor in request.POST.items():
            if campo == "csrfmiddlewaretoken":
                continue

            field = obj._meta.get_field(campo)

            # ForeignKey
            if field.is_relation and hasattr(field, "related_model"):
                try:
                    instance = field.related_model.objects.get(id=valor)
                    setattr(obj, campo, instance)
                except:
                    pass
                continue

            setattr(obj, campo, valor)

        obj.save()

        registrar_actividad(
            request,
            accion="Editar registro",
            modulo="Administración",
            detalle=f"Modelo: {modelo} | ID {obj.id}"
        )

        messages.success(request, "Cambios guardados correctamente")
        return redirect(f"/panel-admin/editar/{modelo}/{obj.id}/")

    # ===================================================
    #           TEMPLATE + DATOS PARA SELECT
    # ===================================================

    if modelo == "pallet":
        template = "packing/admin/panel_admin_editar_pallet.html"
        extra = {
            "productores": Productor.objects.all(),
            "variedades": Variedad.objects.all(),
            "envases": TipoEnvase.objects.all(),
        }

    elif modelo == "terminados":
        template = "packing/admin/panel_admin_editar_terminado.html"
        extra = {
            "distribuidores": Distribuidor.objects.all(),
            "variedades": Variedad.objects.all(),
        }

    elif modelo == "iqf":
        template = "packing/admin/panel_admin_editar_iqf.html"
        extra = {}  # de momento no necesitamos selects

    else:
        template = "packing/admin/panel_admin_editar.html"
        extra = {}

    return render(request, template, {
        "modelo": modelo,
        "obj": obj,
        **extra
    })


#suarios 

@user_passes_test(es_admin_local)
def panel_admin_crear_usuario(request):
    grupos = Group.objects.all()

    if request.method == "POST":
        username = request.POST.get("username")
        correo = request.POST.get("correo")
        password = request.POST.get("password")
        grupo_id = request.POST.get("grupo")

        if User.objects.filter(username=username).exists():
            messages.error(request, "⚠ Este usuario ya existe.")
        else:
            user = User.objects.create_user(
                username=username,
                email=correo,
                password=password
            )
            if grupo_id:
                grupo = Group.objects.get(id=grupo_id)
                user.groups.add(grupo)
            registrar_actividad(
                request,
                accion="Crear usuario",
                modulo="Administración",
                detalle=f"Usuario creado: {username}"
            )

            messages.success(request, "✅ Usuario creado exitosamente.")
            return redirect("panel_admin_usuarios")

    return render(request, "packing/admin/panel_admin_crear_usuario.html", {
        "grupos": grupos
    })


@user_passes_test(es_admin_local)
def panel_admin_editar_usuario(request, id):
    usuario = get_object_or_404(User, id=id)
    grupos = Group.objects.all()

    if request.method == "POST":
        usuario.username = request.POST.get("username")
        usuario.email = request.POST.get("correo")

        nuevo_grupo = request.POST.get("grupo")
        usuario.groups.clear()
        if nuevo_grupo:
            usuario.groups.add(Group.objects.get(id=nuevo_grupo))

        nueva_password = request.POST.get("password")
        if nueva_password:
            usuario.set_password(nueva_password)

        usuario.save()

        registrar_actividad(
            request,
            accion="Editar usuario",
            modulo="Administración",
            detalle=f"Usuario editado: {usuario.username}"
        )
        messages.success(request, "✨ Usuario actualizado correctamente.")
        return redirect("panel_admin_usuarios")

    return render(request, "packing/admin/panel_admin_editar_usuario.html", {
        "usuario": usuario,
        "grupos": grupos
    })


# ➜ ELIMINAR USUARIO
@user_passes_test(es_admin_local)
def panel_admin_eliminar_usuario(request, id):
    usuario = get_object_or_404(User, id=id)

    if not usuario.is_superuser:  # seguridad
        usuario.delete()
        registrar_actividad(
            request,
            accion="Eliminar usuario",
            modulo="Administración",
            detalle=f"Usuario eliminado: {usuario.username}"
        )
        messages.success(request, "🗑 Usuario eliminado correctamente.")

    return redirect("panel_admin_usuarios")




def despachar_iqf_frio(request, id):
    registro = get_object_or_404(IQFDescarte, id=id)

    if request.method == "GET":
        return render(request, "packing/frio/despachar_iqf.html", {
            "registro": registro
        })

    if request.method == "POST":
        peso_final = request.POST.get("peso_final", "").replace(",", ".")

        try:
            peso_final = float(peso_final)
        except:
            messages.error(request, "⚠ Debes ingresar un número válido.")
            return redirect("despachar_iqf_frio", id=id)

        registro.peso_final = peso_final
        registro.estado = "DESPACHADO"
        registro.fecha_retiro = timezone.now()
        registro.save()

        registrar_actividad(
            request,
            accion="Despachar IQF",
            modulo="Frío",
            detalle=f"Grupo {registro.grupo_proceso.id_grupo} | Peso final: {peso_final} kg"
        )

        messages.success(request, "🚚 IQF despachado correctamente.")
        return redirect("lista_iqf_full")



def iqf_dashboard(request):

    pallets_iqf = IQFDescarte.objects.filter(estado="EN_IQF").count()
    retirados = IQFDescarte.objects.filter(estado="RETIRADO").count()
    riesgo = IQFDescarte.objects.filter(
        estado="EN_IQF",
        fecha_ingreso__lte=timezone.now() - timedelta(days=10)
    ).count()

    iqf_recent = IQFDescarte.objects.all().order_by("-id")[:7]
    variedades = IQFDescarte.objects.values_list("variedad", flat=True).distinct()

    return render(request, "packing/control/iqf_dashboard.html", {
        "pallets_iqf": pallets_iqf,
        "retirados": retirados,
        "riesgo": riesgo,
        "iqf_recent": iqf_recent,
        "variedades": variedades,
    })










@login_required
@user_passes_test(es_admin_local)
def panel_admin_actividad(request):
    registros = RegistroActividad.objects.all().order_by('-fecha')

    return render(request, 'packing/admin/actividad.html', {
        'registros': registros
    })


    # 🔍 AUDITORÍA DEL SISTEMA (UTILIDAD GLOBAL)


def registrar_actividad(request, accion, modulo, detalle=""):
    """
    Registra una acción del usuario en la auditoría del sistema
    """
    usuario = request.user if request.user.is_authenticated else None

    RegistroActividad.objects.create(
        usuario=usuario,
        accion=accion,
        modulo=modulo,
        detalle=detalle
    )

    
@login_required
@user_passes_test(es_control)
def control_graficos(request):

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    filtros_pallet = {}
    filtros_terminados = {}

    if fecha_inicio:
        filtros_pallet["fecha_creacion__date__gte"] = fecha_inicio
        filtros_terminados["fecha__date__gte"] = fecha_inicio

    if fecha_fin:
        filtros_pallet["fecha_creacion__date__lte"] = fecha_fin
        filtros_terminados["fecha__date__lte"] = fecha_fin

    pallets_por_dia = (
        Pallet.objects.filter(**filtros_pallet)
        .annotate(dia=TruncDate("fecha_creacion"))
        .values("dia")
        .annotate(total=Count("id"))
        .order_by("dia")
    )

    pallets_por_productor = (
        Pallet.objects.filter(**filtros_pallet)
        .values("productor__nombre")
        .annotate(total=Count("id"))
        .order_by("productor__nombre")
    )

    estados = {
        "procesados": Pallet.objects.filter(procesado=True, **filtros_pallet).count(),
        "no_procesados": Pallet.objects.filter(procesado=False, **filtros_pallet).count(),
        "enviados": PalletTerminado.objects.filter(
            estado="Enviado",
            **filtros_terminados
        ).count(),
    }

    return render(request, "packing/control/graficos.html", {
        "pallets_por_dia": pallets_por_dia,
        "pallets_por_productor": pallets_por_productor,
        "estados": estados,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    })

##a

@login_required
@user_passes_test(lambda u: u.groups.filter(name="Frio").exists())
def lista_iqf_frio(request):
    iqf = IQFDescarte.objects.all().order_by("-id")

    buscar = request.GET.get("buscar", "").strip()
    if buscar:
        iqf = iqf.filter(
            Q(id__icontains=buscar) |
            Q(grupo_proceso__id_grupo__icontains=buscar)
        )

    estado = request.GET.get("estado", "")
    if estado:
        iqf = iqf.filter(estado=estado)

    return render(request, "packing/frio/lista_iqf.html", {
        "iqf": iqf,
        "buscar": buscar,
        "estado_filtro": estado,
    })

##b

@login_required
@user_passes_test(lambda u: u.groups.filter(name="Frio").exists())
def despachar_iqf_frio(request, id):
    registro = get_object_or_404(IQFDescarte, id=id)

    if registro.estado != "EN_IQF":
        messages.error(request, "⚠ Este IQF ya fue despachado.")
        return redirect("lista_iqf_frio")

    if request.method == "GET":
        return render(request, "packing/frio/despachar_iqf.html", {
            "registro": registro
        })

    peso_final = request.POST.get("peso_final", "").replace(",", ".")

    try:
        peso_final = float(peso_final)
    except ValueError:
        messages.error(request, "⚠ Ingresa un peso válido.")
        return redirect("despachar_iqf_frio", id=id)

    registro.peso_final = peso_final
    registro.estado = "DESPACHADO"
    registro.fecha_retiro = timezone.now()
    registro.save()

    registrar_actividad(
        request,
        accion="Despachar IQF",
        modulo="Frío",
        detalle=f"Grupo {registro.grupo_proceso.id_grupo} | Peso final: {peso_final} kg"
    )

    messages.success(request, "🚚 IQF despachado correctamente.")
    return redirect("lista_iqf_frio")
